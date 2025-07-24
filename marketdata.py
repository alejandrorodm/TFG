"""
This file contains the functions to retrieve the data from the KuCoin API.
"""

from datetime import datetime
from kucoin.client import Market
from openpyxl import Workbook
from openpyxl import load_workbook
from datetime import datetime
import pandas as pd
import threading


#  MarketData

"""
    Client to obtain data, we access to kucoin api
"""
client = Market(url='https://api.kucoin.com')

"""
    TIMESTAMP CONVERSION

"""
def getServerTimestamp():
    """Get the server timestamp.

    Returns:
        timestamp(int): timestamp in milliseconds
    """
    return client.get_server_timestamp()

def getServerTime():
    """Get the server time.

    Returns:
        datetime: server time
    """
    return datetime.fromtimestamp(getServerTimestamp()/1000)

def getDatetime(year, month, day, hour, minute, second):
    """Define a date in datetime format.

    Args:
        year (int): year
        month (int): month
        day (int): day
        hour (int): hour
        minute (int): minute
        second (int): second

    Returns:
        datetime: date
    """
    return datetime(year, month, day, hour, minute, second)

def getTimestamp_to_date(timestamp):
    """Convert timestamp (ms) to date (day, month, year).

    Args:
        timestamp (int): timestamp in milliseconds

    Returns:
        datetime: timestamp in date format (day, month, year)
    """
    return datetime.fromtimestamp(int(timestamp)).strftime('%d-%m-%Y')

def getTimestamp_to_datetime(timestamp):
    """Convert timestamp (ms) to date.

    Returns:
        datetime: timestamp in date format
    """
    return datetime.fromtimestamp(int(timestamp/1000))
    
def getDatetime_to_timestamp(date):
    """Convert date to timestamp (ms).
    It returned in ms to be consistent with the server timestamp.
    KuCoin uses ms for the timestamp.
    Returns:
        int: date in timestamp format (ms)
    """
    return int(date.timestamp())

def getNowTimestamp():
    """Get the current timestamp.

    Returns:
        int: current timestamp in milliseconds
    """
    return getDatetime_to_timestamp(datetime.now())

def getNowDate():
    """Get the current date.

    Returns:
        datetime: current date
    """
    return datetime.now()



"""
    UTILS

"""

def getHeader(excelFile):
    """Obtain the header of the Excel file.

    Args:
        excelFile (str): name of the Excel file

    Returns:
        list: header of the Excel file
    """
    wb = load_workbook(excelFile)
    ws = wb.active
    header = []

    for cell in ws[1]:
        header.append(cell.value)

    return header



"""
    WRITE ON EXCEL THE DATA RETRIEVED
"""

def write_to_excel(data, filename):
    """Write the data to an Excel file.

    Args:
        data (list): list of lists containing the data
        filename (str): name of the file
    """
    # Create a new workbook
    wb = Workbook()

    # Select the active sheet
    ws = wb.active

    ws.append(["Timestamp", "Open", "Close", "High", "Low", "Volume"])

    # Empty lists to store the values
    timestamps, open_prices, close_prices, high_prices, low_prices, volumes = ([] for _ in range(6))

    lastrow = 1

    # Iterate over the data and write them to the Excel file
    for sublist in data:
        timestamp, open_price, close_price, high_price, low_price, volume = sublist

        # Add the values to the current row
        row = [timestamp, float(open_price), float(close_price), float(high_price), float(low_price), float(volume)]
        ws.append(row)
        
        lastrow += 1
        timestamps.append(timestamp)
        open_prices.append(float(open_price))
        close_prices.append(float(close_price))
        high_prices.append(float(high_price))
        low_prices.append(float(low_price))
        volumes.append(float(volume))

    ws['E2'] = lastrow
    wb.save(filename)

    print(f"\nData has been written to {filename}")


"""
    GET DATA FROM KUCOIN API
    
"""
    
def getData(coin, interval, start, end, shared_data=None, writeOnExcel=False):
    """Get the data from the KuCoin API.
    Due to limitations in KuCoin API it only can retrieve 1500 elements.
    
    Args:
        coin (str): coin pair (e.g. BTC-USDT)
        interval (str): interval (1min, 5min, 15min, 1hour, 4hour, 8hour, 1day, 1week, 1month)
        start (int): start timestamp
        end (int): end timestamp
        writeOnExcel (bool): write the data to an Excel file (default: False)

    Returns:
        list: data
    """
    data = client.get_kline(coin, interval, startAt=start, endAt=end)
    file = f"{coin}.{interval}.{start}.{end}.xlsx"

    if writeOnExcel:
        write_to_excel(data, file)
    
    if shared_data is not None:
        shared_data.extend(data)

    return data

"""
    MULTI-THREADING

"""

# Temporalidades en diccionario y su calculo
temporalities = {
        '1min': 1500 * 60,
        '5min': 1500 * 60 * 5,
        '15min': 1500 * 60 * 15,
        '1hour': 1500 * 60 * 60,
        '4hour': 1500 * 60 * 60 * 4,
        '8hour': 1500 * 60 * 60 * 8,
        '1day': 1500 * 60 * 60 * 24,
        '1week': 1500 * 60 * 60 * 24 * 7,
        '1month': 1500 * 60 * 60 * 24 * 30
}

import math 

def multi_threading(start_timestamp, end_timestamp, temporality, coin, writeOnExcel=True):
    """Retrieve the data in multiple threads.

    Args:
        start_timestamp (int): start timestamp
        end_timestamp (int): end timestamp
        temporality (str): temporality (1min, 5min, 15min, 1hour, 4hour, 8hour, 1day, 1week, 1month)
        coin (str): coin pair (e.g. BTC-USDT)
        writeOnExcel (bool): decide if you want to write the retrieved data on Excel. Default value: True
    
    Returns:
        list: shared list with the data
        excelFile (str): name of the Excel file where the data is stored
    """

    num_threads = math.ceil((end_timestamp - start_timestamp) / temporalities[temporality])
    
    # if num_threads == 1:
    #     num_threads = 3
    
    # if num_threads == 0:
    #     num_threads = 1
        
    threads = []
    shared_data = []
    lock = threading.Lock()

    def thread_safe_getData(coin, temporality, start, end, shared_data, writeOnExcel):
        try:
            data = getData(coin, temporality, start, end, [], writeOnExcel)
            with lock:
                shared_data.extend(data)
        except Exception as e:
            print(f"Error en hilo para {coin} ({start} - {end}): {e}")


    for i in range(num_threads):
        #print(f"Thread {i+1} of {num_threads}")
        #print("Start timestamp:", start_timestamp + i * temporalities[temporality])
        
        # Calculate the range of timestamps for the current thread
        start = start_timestamp + i * temporalities[temporality]

        # Calculate the end timestamp for the current thread
        end = min(start_timestamp + (i + 1) * temporalities[temporality], end_timestamp)

        thread = threading.Thread(target=thread_safe_getData, args=(coin, temporality, start, end, shared_data, False))
        threads.append(thread)
        thread.start()

    for thread in threads:
        thread.join()

    # Sort the shared data by timestamp (the first attribute) to ensure it is in order
    shared_data.sort(key=lambda x: x[0])

    header = ["Timestamp", "Open", "Close", "High", "Low", "Volume", "base_volume"]
    df = pd.DataFrame(shared_data, columns=header)
    
    if writeOnExcel:
        # Write the data to an Excel file
        file = f'{coin}.{temporality}.{getTimestamp_to_date(start_timestamp)}.{getTimestamp_to_date(end_timestamp)}'
        excelFile = file + '.xlsx'

        # Write the data to an Excel file
        df.to_excel(excelFile, index=False)
        print(f"\nData has been written to {excelFile}")
        
        return df, excelFile
    else:
        return df, None

